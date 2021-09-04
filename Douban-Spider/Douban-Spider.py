# -*- coding: utf-8 -*-
import requests 
from bs4 import BeautifulSoup
import openpyxl
import time 
import re 


# fetchURL: basic function to get HTML
def fetchURL(url): 
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "\
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36 Edg/92.0.902.84"
    }
    r = requests.get(url=url, headers=headers)
    if r.status_code != 200:
        print("Status Code: {}".format(r.status_code))
        return ""
    else:
        html = r.text
        return html 
 

# fetchData: get valuable data 
def fetchData(basic):
    # regex patterns 
    linkPattern = re.compile(r'<a href="(.*?)">') # video detail 
    imgPattern = re.compile(r'<img.*src="(.*?)"', re.S) # img URL 
    titlePattern = re.compile(r'<span class="title">(.*)</span>') # title 
    ratingPattern = re.compile(r'<span class="rating_num" property="v:average">(.*)</span>') # rating 
    commentPattern = re.compile(r'<span>(\d*)人评价</span>') # number of comments 
    overviewPattern = re.compile(r'<span class="inq">(.*)</span>') # overview 
    infoPattern = re.compile(r'<p class="">(.*?)</p>',re.S) # director and more 

    removePattern = re.compile(r'                            |\n|</br>|\.*') # irrelevant content
    brPattern = re.compile(r'<br(\s+)?\/?>(\s+)?') # <br> 

    # empty data list 
    datalist = list()

    # fetch 10 pages 
    for i in range(0, 10):
        url = basic + str(25*i)
        html = fetchURL(url)
        soup = BeautifulSoup(html, "html.parser")
        for item in soup.find_all("div", class_='item'):
            data = list()
            item = str(item)
            # data 
            link = re.findall(linkPattern, item)[0]
            img = re.findall(imgPattern, item)[0]
            data.append(link)
            data.append(img)

            titles= re.findall(titlePattern, item) 
            # check if the movie has both a Chinese title and a foreign language title 
            if(len(titles)>1):
                ctitle = titles[0]
                ftitle = titles[1].replace("/", "")
                data.append(ctitle)
                data.append(ftitle)
            else:
                data.append(titles[0])
                data.append(" ")
            
            rating = re.findall(ratingPattern, item)[0]
            data.append(rating)

            comment = re.findall(commentPattern, item)[0]
            data.append(comment)

            overview = re.findall(overviewPattern, item)
            if len(overview) != 0:
                overview = overview[0].replace("。", '')
                data.append(overview)
            else:
                data.append(" ")
            
            info = re.findall(infoPattern, item)[0]
            info = re.sub(removePattern, "", info)
            info = re.sub(brPattern, " ", info)
            info = re.sub('/', " ", info)
            data.append(info.strip())

            datalist.append(data)
        time.sleep(0.5)
    time.sleep(0.2)
    return datalist


def saveData(datalist, datapath):
    book = openpyxl.Workbook()
    sheet = book.create_sheet("Douban Movie Top 250")
    col = ('URL', "Poster", "Chinese Title", "Foreign Language Title", 
    "Ratings", "Number of Comments", "Overview", "More Info")
    sheet.append(col)
    for i in range(0, 250):
        data = datalist[i]
        for j in range(0, 8):
            sheet.cell(row=(i+2), column=(j+1), value=data[j])
    book.save(datapath)


if __name__ == "__main__":
    print("spider launched...")
    basic = "https://movie.douban.com/top250?start="
    datalist = fetchData(basic)
    datapath = u'./DoubanMovieTop250.xlsx'
    # datapath = u'/home/aistudio/data/DoubanMovieTop250.xlsx'
    saveData(datalist=datalist, datapath=datapath)
    print("finished")